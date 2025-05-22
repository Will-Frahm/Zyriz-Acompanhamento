import pandas as pd
import sqlite3
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from datetime import datetime
import os

# Caminhos
caminho_template = "Templates\\template-acompanhamento.xlsx"
caminho_banco = r"\\wts\instala\Zyriz_relatorio_base\bd_zyriz.db"

# mudar para teste !!!!!!!!!!!!
pasta_saida = r"S:\PSCF\Relatórios PSCF\Monitoramento\Restritos"
# !!!!!!!!!!!
pasta_saida = r"C:\Users\willian\Desktop\Programas\Zyriz Acompanhamento\Emails\Restritos"

os.makedirs(pasta_saida, exist_ok=True)

data_atual = datetime.now().strftime("%d/%m/%Y")

conexao = sqlite3.connect(caminho_banco)

consulta = f'''
    SELECT 
        dados_restritos.cod_produto,
        dados_restritos.produto,
        dados_restritos.link,
        dados_restritos.vendedor,
        dados_restritos.preco,
        dados_restritos.preco_sugerido,
        dados_restritos.diferenca_preco_sugerido,
        dados_restritos.grupo_nome,
        irregularidades_restritos.data_irregular AS primeiro_dia_irregular,     
        irregularidades_restritos.dias_irregular AS dias_consecutivos_irregulares, 
        dados_restritos.id_anuncio
    FROM dados_restritos
    JOIN irregularidades_restritos ON dados_restritos.id_anuncio = irregularidades_restritos.id_anuncio
    WHERE data = '{data_atual}'
'''

df = pd.read_sql(consulta, conexao)
conexao.close()

# Extrair cliente e representante
df["Cliente"] = df["grupo_nome"].str.extract(r"^(.*?)-")
df["Representante"] = df["grupo_nome"].str.extract(r"-(.*)$")
df.drop(columns=["grupo_nome"], inplace=True)

# Função para salvar Excel por cliente
def salvar_relatorio_cliente(df_cliente, cliente_nome, representante_nome):
    if df_cliente.empty:
        return

    CORES_PERIODO = {
        (1, 1): 'e2f2d3',
        (2, 2): 'fcd4b4',
        (3, float('inf')): 'FF7E7E'
    }

    wb = load_workbook(caminho_template)
    ws = wb["Resumo"]

    ws.cell(row=1, column=1, value=f"FRAHM - Produtos RESTRITOS - {data_atual}")
    amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    azul = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    fonte_link = Font(color="0000FF", underline="single")
    linha = 4

    representante_nome = representante_nome or "Desconhecido"

    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=9)
    cell_rep = ws.cell(row=linha, column=1, value=f"Representante: {representante_nome}")
    cell_rep.fill = amarelo
    cell_rep.font = Font(bold=True)
    cell_rep.alignment = Alignment(horizontal="center")
    linha += 1

    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=9)
    cell_cli = ws.cell(row=linha, column=1, value=f"Cliente: {cliente_nome}")
    cell_cli.font = Font(bold=True)
    cell_cli.fill = azul
    cell_cli.alignment = Alignment(horizontal="center")
    linha += 1

    headers = ["codigo", "nome", "link", "vendedor", "preço", "psi", "diferença", 
               "inicio irregular", "dias irregular"]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=linha, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    linha += 1

    for row in df_cliente.itertuples(index=False):
        fill_cor = None
        dias_irreg = row.dias_consecutivos_irregulares
        for (min_d, max_d), cor in CORES_PERIODO.items():
            if min_d <= dias_irreg <= max_d:
                fill_cor = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
                break

        valores = [ 
            row.cod_produto, row.produto, row.link, row.vendedor, row.preco,
            row.preco_sugerido, row.diferenca_preco_sugerido,
            row.primeiro_dia_irregular, row.dias_consecutivos_irregulares 
        ]

        for col, value in enumerate(valores, start=1):
            cell = ws.cell(row=linha, column=col, value=value)

            if col == 1:
                cell.alignment = Alignment(horizontal="center")

            if col == 3:
                cell.hyperlink = value
                cell.font = fonte_link

            if col == 7:
                cell.value = str(round(float(value), 2))

            if fill_cor and col in {8, 9}:
                cell.fill = fill_cor
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

        linha += 1
    linha += 1

    nome_arquivo = f"{cliente_nome.strip().replace('/', '_').replace(' ', '_')}.xlsx"
    caminho_arquivo = os.path.join(pasta_saida, nome_arquivo)
    wb.save(caminho_arquivo)
    print(f"✅ Arquivo salvo: {caminho_arquivo}")

# Agrupar por cliente e exportar
for (cliente, representante), sub_df in df.groupby(["Cliente", "Representante"]):
    salvar_relatorio_cliente(sub_df, cliente, representante)