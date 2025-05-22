import pandas as pd
import sqlite3
from openpyxl import load_workbook
from datetime import datetime

ids_para_remover = ['MLB4523378420','MLB4523162390']  # IDS Bugados da zyriz aqui !!

caminho_banco = r"\\wts\instala\Zyriz_relatorio_base\bd_zyriz.db"
conexao = sqlite3.connect(caminho_banco)

consulta = '''
        WITH DiasIrregulares AS (
            SELECT id_anuncio, data, cod_produto, produto, preco, preco_sugerido, diferenca_preco_sugerido, 
                   link, vendedor, grupo_nome,
                   CASE WHEN preco < preco_sugerido THEN 1 ELSE 0 END AS irregular
            FROM dados_marketplace
        ),
        PreencheGrupoNome AS (
            SELECT id_anuncio, data, cod_produto, produto, preco, preco_sugerido, diferenca_preco_sugerido, 
                   link, vendedor,
                   (SELECT grupo_nome
                    FROM DiasIrregulares d2
                    WHERE d2.vendedor = DiasIrregulares.vendedor
                      AND d2.grupo_nome IS NOT NULL
                      AND d2.data <= DiasIrregulares.data
                    ORDER BY d2.data DESC
                    LIMIT 1) AS grupo_nome_preenchido,
                   irregular
            FROM DiasIrregulares
        ),
        SequenciaIrregular AS (
            SELECT id_anuncio, data, cod_produto, produto, preco, preco_sugerido, diferenca_preco_sugerido, 
                   link, vendedor, grupo_nome_preenchido, irregular,
                   SUM(CASE WHEN irregular = 0 THEN 1 ELSE 0 END) OVER (PARTITION BY id_anuncio ORDER BY data) AS grupo
            FROM PreencheGrupoNome
        ),
        SequenciaAgrupada AS (
            SELECT id_anuncio, data, cod_produto, produto, preco, preco_sugerido, diferenca_preco_sugerido, 
                   link, vendedor, grupo_nome_preenchido, irregular, grupo
            FROM SequenciaIrregular
            WHERE irregular = 1  
            AND diferenca_preco_sugerido < 0  
        )
        SELECT cod_produto, produto, preco, preco_sugerido, diferenca_preco_sugerido,
               link, vendedor, 
               CASE 
                   WHEN INSTR(grupo_nome_preenchido, '-') > 0 
                   THEN SUBSTR(grupo_nome_preenchido, 1, INSTR(grupo_nome_preenchido, '-') - 1)
                   ELSE grupo_nome_preenchido
               END AS Cliente,
               CASE 
                   WHEN INSTR(grupo_nome_preenchido, '-') > 0 
                   THEN SUBSTR(grupo_nome_preenchido, INSTR(grupo_nome_preenchido, '-') + 1)
                   ELSE NULL
               END AS Representante,
               MIN(data) AS primeiro_dia_irregular, MAX(data) AS ultimo_dia_irregular, 
               COUNT(*) AS dias_consecutivos_irregulares, id_anuncio
        FROM SequenciaAgrupada
        GROUP BY cod_produto, produto, preco, preco_sugerido, diferenca_preco_sugerido, 
                 link, vendedor, grupo_nome_preenchido, id_anuncio, grupo
        ORDER BY produto, grupo_nome_preenchido, vendedor;
'''

df = pd.read_sql(consulta, conexao)
conexao.close()

df["primeiro_dia_irregular"] = pd.to_datetime(df["primeiro_dia_irregular"], dayfirst=True, errors="coerce").dt.strftime("%d/%m/%Y")
df["ultimo_dia_irregular"] = pd.to_datetime(df["ultimo_dia_irregular"], dayfirst=True, errors="coerce").dt.strftime("%d/%m/%Y")

# filtro
df_restritos = df[df['produto'].str.startswith("(RESTRITOS)", na=False)]
df_restritos = df_restritos[~df_restritos['produto'].str.startswith("(NÃO MONITORADO)", na=False)]

df_nao_restritos = df[~df['produto'].str.startswith("(RESTRITOS)", na=False)]
df_nao_restritos = df_nao_restritos[~df_nao_restritos['produto'].str.startswith("(NÃO MONITORADO)", na=False)]

df_restritos = df_restritos[~df_restritos["id_anuncio"].isin(ids_para_remover)]
df_nao_restritos = df_nao_restritos[~df_nao_restritos["id_anuncio"].isin(ids_para_remover)]

caminho_template = "Templates\\template-acompanhamento.xlsx"
caminho_saida_restritos = "relatorio_restritos.xlsx"
caminho_saida_nao_restritos = "relatorio_nao_restritos.xlsx"

def salvar_relatorio(df, caminho_saida):
    if df.empty:
        print(f"❌ Nenhum dado para {caminho_saida}. Arquivo não gerado.")
        return
    
    wb = load_workbook(caminho_template)

    # planilha detalhada
    ws = wb["DETALHADO"]        
    data_atual = datetime.now().strftime("%d/%m/%Y")
    ws.cell(row=1, column=1, value=f"FRAHM - Monitoramento de PSI - {data_atual}")

    for r_idx, row in enumerate(df.itertuples(index=False), start=3):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(caminho_saida)
    print(f"✅ Relatório salvo: {caminho_saida}")

    # planilha detalhada
    ws = wb["Resumo"]

salvar_relatorio(df_restritos, caminho_saida_restritos)
salvar_relatorio(df_nao_restritos, caminho_saida_nao_restritos) 