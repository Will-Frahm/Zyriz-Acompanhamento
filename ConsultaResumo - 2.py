import pandas as pd
import sqlite3
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from datetime import datetime

# Caminhos
caminho_template = "Templates\\template-acompanhamento.xlsx"
caminho_saida = "relatorio_final.xlsx"
caminho_banco = r"\\wts\instala\Zyriz_relatorio_base\bd_zyriz.db"
ids_para_remover = ['']

data_atual = datetime.now().strftime("%d/%m/%Y")
# Conexão e SQL
conexao = sqlite3.connect(caminho_banco)

consulta = f'''
    SELECT 
        dados_marketplace.cod_produto,
        dados_marketplace.produto,
        dados_marketplace.link,
        dados_marketplace.vendedor,
        dados_marketplace.preco,
        dados_marketplace.preco_sugerido,
        dados_marketplace.diferenca_preco_sugerido,
        dados_marketplace.grupo_nome,
        irregularidades.data_irregular AS primeiro_dia_irregular,     
        irregularidades.dias_irregular AS dias_consecutivos_irregulares, 
        dados_marketplace.id_anuncio
    FROM dados_marketplace
    JOIN irregularidades ON dados_marketplace.id_anuncio = irregularidades.id_anuncio
    WHERE data = '{data_atual}'
'''

print(data_atual)

df = pd.read_sql(consulta, conexao)
conexao.close()

# Separar grupo_nome em Cliente e Representante
df["Cliente"] = df["grupo_nome"].str.extract(r"^(.*?)-")
df["Representante"] = df["grupo_nome"].str.extract(r"-(.*)$")
df.drop(columns=["grupo_nome"], inplace=True)

df = df[~df["produto"].str.startswith("(RESTRITOS)")]
df = df[~df["produto"].str.startswith("(NÃO MONITORADO)")]

# Função de exportação
def salvar_relatorio(df, caminho_saida):
    if df.empty:
        print(f"❌ Nenhum dado disponível. Arquivo não gerado.")
        return        

    CORES_PERIODO = {
        (1, 1): 'e2f2d3',    # Verde claro
        (2, 2): 'fcd4b4',   # Laranja claro
        (3, float('inf')): 'FF7E7E'  # Vermelho claro
    }
    
    wb = load_workbook(caminho_template)
    ws = wb["Resumo"]
        
    ws.cell(row=1, column=1, value=f"FRAHM - Monitoramento de PSI - {data_atual}")
    
    amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    azul = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    fonte_link = Font(color="0000FF", underline="single")
    linha = 4

    for (rep, cli), sub_df in df.groupby(["Representante", "Cliente"]):  
        rep = rep if pd.notna(rep) else "Desconhecido"          
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=9)
        cell_rep = ws.cell(row=linha, column=1, value=f"Representante: {rep}")
        cell_rep.fill = amarelo
        cell_rep.font = Font(bold=True)
        cell_rep.alignment = Alignment(horizontal="center")
        linha += 1
        
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=9)
        cell_cli = ws.cell(row=linha, column=1, value=f"Cliente: {cli}")
        cell_cli.font = Font(bold=True)
        cell_cli.fill = azul
        cell_cli.alignment = Alignment(horizontal="center")
        linha += 1    
            
        headers = ["codigo", "nome", "link", "vendedor", "preço", "psi", "diferença", 
                   "inicio irregular", "dias irregular"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=linha, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        linha += 1
        
        for row in sub_df.itertuples(index=False):
            fill_cor = None
            dias_irreg = row.dias_consecutivos_irregulares
            for (min_d, max_d), cor in CORES_PERIODO.items():
                if min_d <= dias_irreg <= max_d:
                    fill_cor = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
                    break
                        
            valores = [ 
                row.cod_produto,       
                row.produto,           
                row.link,              
                row.vendedor,          
                row.preco,            
                row.preco_sugerido,    
                row.diferenca_preco_sugerido,
                row.primeiro_dia_irregular,  
                row.dias_consecutivos_irregulares 
            ]
            
            for col, value in enumerate(valores, start=1):
                cell = ws.cell(row=linha, column=col, value=value)                
            
                if col == 1:
                    cell.alignment = Alignment(horizontal="center") 

                if col == 3:
                    cell.hyperlink = value
                    cell.font = fonte_link

                if col == 7:
                    cell.value = str(round(float(value), 2))                    

                if fill_cor and col in {8, 9}:
                    cell.fill = fill_cor
                    cell.font = Font(bold=True) 
                    cell.alignment = Alignment(horizontal="center") 
            
            linha += 1
        linha += 1

    wb.save(caminho_saida)
    print(f"✅ Relatório salvo: {caminho_saida}")
    
# Executar
salvar_relatorio(df, caminho_saida)