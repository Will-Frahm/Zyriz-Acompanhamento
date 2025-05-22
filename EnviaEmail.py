import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
from openpyxl import load_workbook
from datetime import datetime

class OutlookEmail:
    def __init__(self, remetente, senha):
        self.remetente = remetente
        self.senha = senha
        self.smtp_server = "smtp.gmail.com"
        self.port = 587

    def pegar_nome_cliente(self, caminho_arquivo):
        wb = load_workbook(caminho_arquivo, data_only=True)
        ws = wb["Resumo"]
        valor = ws["A5"].value 
        if valor and "Cliente:" in valor:
            return valor.split("Cliente:")[-1].strip()
        return None

    def enviar_email(self, destinatarios, assunto, corpo):
        try:
            for destinatario, info in destinatarios.items():
                caminho_anexo = info.get("caminho")
                cc_list = info.get("cc", [])

                if not caminho_anexo or not os.path.exists(caminho_anexo):
                    print(f"⚠️ Anexo não encontrado para {destinatario}. E-mail não será enviado.")
                    continue

                # Pegar o nome do cliente a partir do arquivo
                nome_cliente = self.pegar_nome_cliente(caminho_anexo)
                if nome_cliente:
                    corpo_email = corpo.format(nome_cliente)

                msg = MIMEMultipart()
                msg['From'] = self.remetente
                msg['To'] = destinatario
                msg['Cc'] = ", ".join(cc_list)
                msg['Subject'] = assunto
                msg.attach(MIMEText(corpo_email, 'html'))

                with open(caminho_anexo, "rb") as anexo:
                    parte = MIMEBase('application', 'octet-stream')
                    parte.set_payload(anexo.read())
                    encoders.encode_base64(parte)
                    parte.add_header(   
                        'Content-Disposition',
                        f'attachment; filename={os.path.basename(caminho_anexo)}'
                    )
                    msg.attach(parte)

                with smtplib.SMTP(self.smtp_server, self.port) as servidor:
                    servidor.starttls()
                    servidor.login(self.remetente, self.senha)
                    servidor.sendmail(
                        self.remetente,
                        [destinatario] + cc_list,
                        msg.as_string()
                    )

                print(f"✅ E-mail enviado com sucesso para {destinatario}!")

        except Exception as e:
            print(f"❌ Erro ao enviar e-mail: {e}")


# --- CONFIGURAÇÃO ---
outlook = OutlookEmail(
    remetente="ouvidoria.mercado@audiofrahm.com.br",
    senha=""
)

# Lista de e-mails e arquivos
destinatarios = {
    "ti@audiofrahm.com.br": {
        "caminho": r"Emails\HZ_som.xlsx",
        "cc": [""]
    }
    
}

# Corpo do e-mail
corpo_email_template = """\
E-MAIL AUTOMÁTICO — FAVOR NÃO RESPONDER <br>
<br>
Olá, {0}!<br>
Identificamos que um ou mais de seus anúncios estão em desacordo com as políticas comerciais da FRAHM. As irregularidades referem-se, principalmente, <br>
à divulgação de produtos <b>restritos para venda em plataformas digitais</b> e/ou à prática de <b>preços abaixo do Preço Mínimo de Anúncio (PMA)</b>, estes considerados predatórios.<br>
Lembramos que é essencial que todos os anúncios estejam alinhados as diretrizes <b>de seu canal de vendas</b>, garantindo uma atuação justa, transparente e saudável no mercado.<br>
<br>
Quando identificamos irregularidades:<br>
<li>
    <ul><b>1º comunicado:</b> Solicitamos que os anúncios em desacordo sejam corrigidos no prazo de até <b>48 horas</b>. (este e-mail é o 1° aviso)</ul>
    <ul><b>2º comunicado:</b> Se os ajustes não forem realizados, uma nova notificação será enviada e as condições comerciais previamente <br>
                              concedidas serão <b>suspensas</b>, tanto nas negociações em andamento quanto nas futuras.<br></ul>
</li>           
<br>
    Para dúvidas ou esclarecimentos, nossa equipe de representantes comerciais está à disposição.<br>
<br>
Confira nos links abaixo:<br>
<li>
    <ul> <b>A lista atualizada de produtos restritos à venda em e-commerce</b><br></ul>
    <ul> <b>A tabela vigente com os valores de Preço Mínimo de Anúncio (PMA)</b></ul>
</li>    
<br>
    https://drive.google.com/drive/folders/1FyLSL38wb82lsi6akdxZE-OqV5i-ODVY<br>
<br>
Agradecemos a compreensão e a parceria de sempre para mantermos um relacionamento sustentável e alinhado às boas práticas comerciais.<br>
Atenciosamente,<br>
<b>Equipe FRAHM</b><br>
"""
data = datetime.now().strftime("%d/%m/%Y")

# Envio
outlook.enviar_email(
    destinatarios=destinatarios,
    assunto=f'[FRAHM] Notificação - Produtos Restritos ({data})',
    corpo=corpo_email_template
)